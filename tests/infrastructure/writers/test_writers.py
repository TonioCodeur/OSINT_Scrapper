"""The five export formats, their schemas and the formula-injection guard (SPEC 9)."""

from __future__ import annotations

import csv
import json
from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from osint_scrapper import SCHEMA_VERSION
from osint_scrapper.domain.attributes import ExtractionLayer, FieldName, Finding, Provenance
from osint_scrapper.domain.crawl import CrawlOutcome, PageOutcome, PageStatus
from osint_scrapper.domain.report import SiteReport
from osint_scrapper.domain.target import CrawlSettings, Purpose, PurposeCategory, resolve_target
from osint_scrapper.infrastructure.writers.csv_writer import CsvReportWriter
from osint_scrapper.infrastructure.writers.json_reader import JsonReportLoader
from osint_scrapper.infrastructure.writers.json_writer import JsonReportWriter
from osint_scrapper.infrastructure.writers.jsonl_writer import JsonlReportWriter
from osint_scrapper.infrastructure.writers.markdown_writer import MarkdownReportWriter
from osint_scrapper.infrastructure.writers.rows import FINDING_COLUMNS, PAGE_COLUMNS
from osint_scrapper.infrastructure.writers.sanitize import FORMULA_TRIGGERS, guard
from osint_scrapper.infrastructure.writers.xlsx_writer import XlsxReportWriter

MOMENT = datetime(2026, 7, 27, 14, 2, 11, tzinfo=UTC)

JSON_SCHEMA = {
    "schema_version",
    "run",
    "target",
    "settings",
    "statistics",
    "findings",
    "pages",
}
RUN_KEYS = {
    "run_id",
    "started_at",
    "finished_at",
    "outcome",
    "outcome_detail",
    "purpose_category",
    "purpose_note",
    "retention_days",
    "tool",
}
FINDING_KEYS = {
    "field",
    "value",
    "extraction_confidence",
    "page_support",
    "occurrence_count",
    "first_seen_url",
    "metadata",
    "provenance",
}


def provenance(url: str = "https://example.com/", raw: str = "contact@example.com") -> Provenance:
    return Provenance(
        source_url=url,
        collected_at=MOMENT,
        extraction_layer=ExtractionLayer.STRUCTURED_DATA,
        raw_value=raw,
    )


def report(**overrides: object) -> SiteReport:
    findings = overrides.pop(
        "findings",
        (
            Finding(
                field=FieldName.EMAIL,
                value="contact@example.com",
                extraction_confidence=0.9,
                page_support=2,
                occurrence_count=3,
                first_seen_url="https://example.com/",
                metadata={"email_kind": "role"},
                provenance=(provenance(),),
            ),
        ),
    )
    defaults: dict[str, object] = {
        "run_id": "r-testrun1",
        "target": resolve_target("example.com", include_subdomains=True),
        "settings": CrawlSettings(),
        "purpose": Purpose(category=PurposeCategory.DUE_DILIGENCE, note="vendor review"),
        "started_at": MOMENT,
        "finished_at": MOMENT,
        "outcome": CrawlOutcome.COMPLETED,
        "outcome_detail": None,
        "tool_name": "osint-scrapper",
        "tool_version": "0.2.0",
        "user_agent": "OSINT-Scrapper/0.2.0 (+https://example.org; contact: a@b.example)",
        "findings": findings,
        "pages": (
            PageOutcome("https://example.com/", 0, PageStatus.OK, None, 200, "text/html", 1),
            PageOutcome(
                "https://example.com/private/",
                1,
                PageStatus.SKIPPED_ROBOTS,
                "robots_disallow",
            ),
        ),
        "pages_fetched": 1,
        "requests_made": 3,
    }
    defaults.update(overrides)
    return SiteReport(**defaults)  # type: ignore[arg-type]


def test_the_json_document_carries_no_key_outside_the_schema(tmp_path: Path) -> None:
    """AC-EXPORT-4: data minimization stays verifiable."""
    JsonReportWriter().write(report(), tmp_path)
    document = json.loads((tmp_path / "report.json").read_text(encoding="utf-8"))
    assert set(document) == JSON_SCHEMA
    assert document["schema_version"] == SCHEMA_VERSION
    assert set(document["run"]) == RUN_KEYS
    assert set(document["findings"][0]) == FINDING_KEYS
    assert "conflicts" not in document


def test_no_page_body_or_free_text_reaches_disk(tmp_path: Path) -> None:
    """AC-EXPORT-4 and SPEC FR-15: only the short raw value survives."""
    JsonReportWriter().write(report(), tmp_path)
    written = (tmp_path / "report.json").read_text(encoding="utf-8")
    assert "<html" not in written
    assert "<body" not in written


def test_json_key_order_is_the_documented_one(tmp_path: Path) -> None:
    """SPEC 9.1: the order is written deliberately, not sorted."""
    JsonReportWriter().write(report(), tmp_path)
    document = json.loads((tmp_path / "report.json").read_text(encoding="utf-8"))
    assert list(document) == [
        "schema_version",
        "run",
        "target",
        "settings",
        "statistics",
        "findings",
        "pages",
    ]


def test_csv_columns_encoding_and_line_endings_match_the_specification(tmp_path: Path) -> None:
    """AC-EXPORT-1."""
    CsvReportWriter().write(report(), tmp_path)
    raw = (tmp_path / "report.csv").read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf"), "Excel needs the BOM to read accented values"
    assert b"\r\n" in raw
    with (tmp_path / "report.csv").open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    assert tuple(rows[0]) == FINDING_COLUMNS
    with (tmp_path / "report_pages.csv").open(encoding="utf-8-sig", newline="") as handle:
        assert tuple(next(csv.reader(handle))) == PAGE_COLUMNS


def test_a_value_a_spreadsheet_would_evaluate_is_neutralized(tmp_path: Path) -> None:
    """AC-EXPORT-1: these files carry text scraped from third-party pages."""
    dangerous = Finding(
        field=FieldName.TECHNOLOGY,
        value="=cmd()",
        extraction_confidence=0.75,
        page_support=1,
        occurrence_count=1,
        first_seen_url="https://example.com/",
        provenance=(provenance(raw="=cmd()"),),
    )
    CsvReportWriter().write(report(findings=(dangerous,)), tmp_path)
    body = (tmp_path / "report.csv").read_text(encoding="utf-8-sig")
    assert "'=cmd()" in body


@pytest.mark.parametrize("trigger", FORMULA_TRIGGERS)
def test_every_documented_trigger_is_guarded(trigger: str) -> None:
    assert guard(f"{trigger}value").startswith("'")


def test_a_value_no_spreadsheet_would_evaluate_is_left_alone() -> None:
    """The guard must not rewrite ordinary scraped text, whatever it starts with."""
    assert guard("plain") == "plain"
    assert guard("Contact us") == "Contact us"
    assert guard("") == ""


def test_the_workbook_has_exactly_four_sheets_with_matching_headers(tmp_path: Path) -> None:
    """AC-EXPORT-2."""
    CsvReportWriter().write(report(), tmp_path)
    XlsxReportWriter().write(report(), tmp_path)
    workbook = load_workbook(tmp_path / "report.xlsx")
    assert workbook.sheetnames == ["Run", "Findings", "Pages", "Compliance"]

    with (tmp_path / "report.csv").open(encoding="utf-8-sig", newline="") as handle:
        csv_header = next(csv.reader(handle))
    findings = workbook["Findings"]
    assert [cell.value for cell in next(findings.iter_rows())] == csv_header

    with (tmp_path / "report_pages.csv").open(encoding="utf-8-sig", newline="") as handle:
        pages_header = next(csv.reader(handle))
    pages = workbook["Pages"]
    assert [cell.value for cell in next(pages.iter_rows())] == pages_header


def test_the_compliance_sheet_states_the_posture_of_the_run(tmp_path: Path) -> None:
    """SPEC 9.3: an auditor can read it without parsing JSON."""
    XlsxReportWriter().write(report(), tmp_path)
    sheet = load_workbook(tmp_path / "report.xlsx")["Compliance"]
    values = {row[0]: row[1] for row in sheet.iter_rows(min_row=2, values_only=True)}
    assert values["robots_txt_honored"] == "true"
    assert values["hard_floor_seconds"] == 0.5
    assert values["pages_skipped_by_robots"] == 1
    assert values["purpose_category"] == "due_diligence"


def test_jsonl_is_one_object_per_line_with_lf_and_no_bom(tmp_path: Path) -> None:
    """AC-EXPORT-3."""
    JsonlReportWriter().write(report(), tmp_path)
    raw = (tmp_path / "report.jsonl").read_bytes()
    assert not raw.startswith(b"\xef\xbb\xbf")
    assert b"\r\n" not in raw
    lines = raw.decode("utf-8").splitlines()
    assert len(lines) == 1
    record = json.loads(lines[0])
    assert record["purpose_category"] == "due_diligence"
    assert record["metadata"] == {"email_kind": "role"}
    assert record["value"] == "contact@example.com"


def test_jsonl_applies_no_formula_guard(tmp_path: Path) -> None:
    """AC-EXPORT-3: mangling values with an apostrophe would corrupt a machine format."""
    dangerous = Finding(
        field=FieldName.TECHNOLOGY,
        value="=cmd()",
        extraction_confidence=0.75,
        page_support=1,
        occurrence_count=1,
        first_seen_url="https://example.com/",
        provenance=(provenance(raw="=cmd()"),),
    )
    JsonlReportWriter().write(report(findings=(dangerous,)), tmp_path)
    record = json.loads((tmp_path / "report.jsonl").read_text(encoding="utf-8").splitlines()[0])
    assert record["value"] == "=cmd()"


def test_markdown_emits_no_raw_html_and_escapes_table_characters(tmp_path: Path) -> None:
    """AC-EXPORT-3: a renderer must not be able to execute what a site published."""
    hostile = Finding(
        field=FieldName.ORGANIZATION_NAME,
        value="<script>alert(1)</script> | `x`",
        extraction_confidence=0.9,
        page_support=1,
        occurrence_count=1,
        first_seen_url="https://example.com/",
        provenance=(provenance(raw="x"),),
    )
    MarkdownReportWriter().write(report(findings=(hostile,)), tmp_path)
    body = (tmp_path / "report.md").read_text(encoding="utf-8")
    assert "<script>" not in body
    assert "\\|" in body
    assert "\\`" in body
    assert "## Compliance" in body
    assert "robots.txt honored" in body


def test_a_written_report_loads_back_identically(tmp_path: Path) -> None:
    """AC-EXPORT-6 depends on this: re-export must not need the network."""
    original = report()
    JsonReportWriter().write(original, tmp_path)
    loaded = JsonReportLoader().load(tmp_path / "report.json")
    assert loaded == original


def test_a_report_from_another_schema_version_is_refused(tmp_path: Path) -> None:
    """A silent partial load would produce an export that lies about its shape."""
    from osint_scrapper.application.errors import ConfigurationError

    path = tmp_path / "report.json"
    path.write_text(json.dumps({"schema_version": "1.0"}), encoding="utf-8")
    with pytest.raises(ConfigurationError):
        JsonReportLoader().load(path)
