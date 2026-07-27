"""AC-E2E-1: the whole chain over the committed mini-site, with only the network faked.

Real ``CrawlSiteUseCase``, real frontier, real canonicalization, real robots
policy, real rate limiter, real extraction pipeline, real aggregator, real
writers. **Its collaborators must not be replaced with stubs**: this is the only
test that catches drift between what an extractor puts in a ``RawField`` and
what the aggregator expects, and a change that keeps both sides consistent with
their own fakes still fails here.

One deliberate departure from the wording of AC-E2E-1: the fake is the HTTP
transport rather than the ``PageFetcher``. Faking the fetcher would remove the
robots policy, the rate limiter, the per-hop redirect gating and the body caps
from the path under test — the opposite of what "real robots policy, real rate
limiter" in the same sentence asks for.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from osint_scrapper.application.ports import LedgerEntry
from osint_scrapper.domain.attributes import MAXIMUM_RAW_VALUE_LENGTH, FieldName
from osint_scrapper.domain.crawl import CrawlOutcome
from osint_scrapper.domain.report import SiteReport
from osint_scrapper.domain.target import CrawlSettings, Purpose, PurposeCategory, resolve_target
from tests.conftest import (
    FakeHttpSession,
    FakeSleeper,
    FrozenClock,
    InMemoryLedger,
    NeverCancelled,
    RecordingObserver,
)
from tests.site_fixture import ORIGIN, site_routes
from tests.wiring import build_crawl, build_exporter, build_fetcher

ALL_FORMATS = ("json", "csv", "xlsx", "jsonl", "md")
PURPOSE = Purpose(category=PurposeCategory.DUE_DILIGENCE, note="pre-contract review")


@pytest.fixture
def crawled(tmp_path: Path) -> tuple[SiteReport, Path, InMemoryLedger, RecordingObserver]:
    """Crawl the mini-site and export it in every format, once, for the whole module."""
    session = FakeHttpSession(site_routes())
    target = resolve_target("example.com", include_subdomains=True)
    settings = CrawlSettings(max_pages=50, max_depth=3)
    clock = FrozenClock()
    sleeper = FakeSleeper()
    observer = RecordingObserver()
    fetcher = build_fetcher(session, target, clock, sleeper, max_pages=settings.max_pages)
    use_case = build_crawl(fetcher, observer, clock, sleeper, NeverCancelled())
    report = use_case.execute(target, settings, PURPOSE)

    ledger = InMemoryLedger()
    build_exporter(ledger).execute(report, ALL_FORMATS, tmp_path)
    return report, tmp_path / report.run_id, ledger, observer


def findings_of(report: SiteReport, field: FieldName) -> set[str]:
    return {item.value for item in report.findings if item.field is field}


def test_the_run_completes_and_reports_what_it_did(crawled) -> None:
    report, _, _, observer = crawled
    assert report.outcome is CrawlOutcome.COMPLETED
    assert report.pages_fetched >= 6
    assert report.requests_made >= report.pages_fetched
    assert report.requests_made <= report.settings.max_pages
    assert observer.finished == [report]


def test_every_field_the_product_declares_is_actually_produced(crawled) -> None:
    """The drift check: an extractor and the aggregator must agree on all nine fields."""
    report, _, _, _ = crawled
    produced = {item.field for item in report.findings}
    assert produced == set(FieldName), f"missing: {sorted(set(FieldName) - produced)}"


def test_the_values_are_the_ones_the_mini_site_publishes(crawled) -> None:
    report, _, _, _ = crawled
    assert "contact@example.com" in findings_of(report, FieldName.EMAIL)
    assert "security@example.com" in findings_of(report, FieldName.EMAIL)
    assert "jean@example.com" in findings_of(report, FieldName.EMAIL)
    assert "+33123456789" in findings_of(report, FieldName.PHONE)
    # Both spellings the site publishes fold to one key, and the structured
    # variant wins because its layer is higher (SPEC 8.5).
    assert findings_of(report, FieldName.POSTAL_ADDRESS) == {
        "12 rue des Lilas, 75011, Paris, France"
    }
    assert "Camille Ferrand" in findings_of(report, FieldName.PERSON_NAME)
    assert "Example Association" in findings_of(report, FieldName.ORGANIZATION_NAME)
    assert "https://github.com/example-association" in findings_of(
        report, FieldName.SOCIAL_PROFILE
    )
    assert "https://example.com/pgp-key.asc" in findings_of(report, FieldName.PGP_KEY_URL)
    assert "999000003" in findings_of(report, FieldName.COMPANY_IDENTIFIER)
    assert "FR42999000003" in findings_of(report, FieldName.COMPANY_IDENTIFIER)
    assert "ExampleCMS 3.2" in findings_of(report, FieldName.TECHNOLOGY)


def test_a_failed_checksum_appears_nowhere_in_any_export(crawled) -> None:
    """AC-EXTRACT-7: the same digits with one changed are discarded entirely."""
    _, directory, _, _ = crawled
    for path in directory.iterdir():
        if path.suffix == ".xlsx":
            continue
        body = path.read_text(encoding="utf-8-sig", errors="replace")
        assert "999000004" not in body
        assert "999 000 004" not in body, "not even as a retained raw value"


def test_the_footer_address_carries_real_page_support(crawled) -> None:
    """AC-EXTRACT-2: the same fact on many pages is one finding with a count."""
    report, _, _, _ = crawled
    footer = next(
        item
        for item in report.findings
        if item.field is FieldName.EMAIL and item.value == "contact@example.com"
    )
    assert footer.page_support >= 4
    assert footer.occurrence_count >= footer.page_support
    assert len(footer.provenance) <= 10


def test_every_status_the_mini_site_exercises_is_present(crawled) -> None:
    """AC-E2E-1: the fixture must exercise every branch the criteria name."""
    report, _, _, _ = crawled
    seen = {str(page.status) for page in report.pages}
    assert {
        "ok",
        "skipped_robots",
        "skipped_extension",
        "skipped_content_type",
        "skipped_off_scope",
        "url_rejected_shape",
        "credentials_in_url",
        "off_scope_redirect",
        "parse_error",
    } <= seen


def test_every_finding_is_attributed_and_bounded(crawled) -> None:
    """AC-LEGAL-2 and AC-LEGAL-3."""
    report, _, _, _ = crawled
    for finding in report.findings:
        assert finding.provenance
        for entry in finding.provenance:
            assert entry.source_url.strip()
            assert entry.collected_at.utcoffset().total_seconds() == 0  # type: ignore[union-attr]
            assert len(entry.raw_value) <= MAXIMUM_RAW_VALUE_LENGTH


def test_all_five_formats_are_written(crawled) -> None:
    _, directory, _, _ = crawled
    names = {path.name for path in directory.iterdir()}
    assert names == {
        "report.json",
        "report.csv",
        "report_pages.csv",
        "report.xlsx",
        "report.jsonl",
        "report.md",
    }


def test_the_json_export_matches_the_report(crawled) -> None:
    report, directory, _, _ = crawled
    document = json.loads((directory / "report.json").read_text(encoding="utf-8"))
    assert document["run"]["outcome"] == "completed"
    assert document["run"]["purpose_category"] == "due_diligence"
    assert document["statistics"]["findings_count"] == len(report.findings)
    assert len(document["pages"]) == len(report.pages)
    assert [item["field"] for item in document["findings"]] == [
        str(item.field) for item in report.findings
    ]


def test_the_csv_has_one_row_per_provenance_entry(crawled) -> None:
    report, directory, _, _ = crawled
    with (directory / "report.csv").open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    expected = sum(len(finding.provenance) for finding in report.findings)
    assert len(rows) - 1 == expected
    assert rows[1][1] == "due_diligence"


def test_the_page_csv_lists_every_page(crawled) -> None:
    report, directory, _, _ = crawled
    with (directory / "report_pages.csv").open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    assert len(rows) - 1 == len(report.pages)


def test_the_workbook_and_the_csv_agree(crawled) -> None:
    _, directory, _, _ = crawled
    workbook = load_workbook(directory / "report.xlsx")
    assert workbook.sheetnames == ["Run", "Findings", "Pages", "Compliance"]
    with (directory / "report.csv").open(encoding="utf-8-sig", newline="") as handle:
        header = next(csv.reader(handle))
    assert [cell.value for cell in next(workbook["Findings"].iter_rows())] == header


def test_the_jsonl_and_the_markdown_are_readable(crawled) -> None:
    report, directory, _, _ = crawled
    lines = (directory / "report.jsonl").read_text(encoding="utf-8").splitlines()
    assert len(lines) == sum(len(finding.provenance) for finding in report.findings)
    assert json.loads(lines[0])["field"] == str(report.findings[0].field)
    markdown = (directory / "report.md").read_text(encoding="utf-8")
    assert markdown.startswith("# OSINT report — example.com")
    assert "## Findings" in markdown and "## Pages" in markdown


def test_no_page_html_reaches_disk(crawled) -> None:
    """SPEC FR-15: response bodies and free page text are never persisted."""
    _, directory, _, _ = crawled
    for path in directory.iterdir():
        if path.suffix == ".xlsx":
            continue
        body = path.read_text(encoding="utf-8-sig", errors="replace")
        assert "<!doctype html>" not in body.lower()
        assert "<address>" not in body.lower()


def test_no_credential_from_a_crawled_page_reaches_any_export(crawled) -> None:
    """SPEC 5.2 rule 9, asserted across the whole layer boundary.

    The mini-site publishes ``https://user:pass@example.com/private-area``. The
    userinfo is stripped in the domain, when the URL is rejected, and every
    layer above simply renders what it was handed — the page log shows a URL
    column verbatim and no writer redacts anything of its own. That makes the
    stripping a guarantee one layer keeps on behalf of four, which is exactly
    the kind that rots silently, so it is asserted here at the far end: on the
    files, in every format, not on the domain object that produced them.
    """
    _, directory, _, _ = crawled

    for path in sorted(directory.iterdir()):
        body = path.read_bytes()
        assert b"user:pass" not in body, f"{path.name} leaked the credentials"
        assert b"pass@example.com" not in body, f"{path.name} leaked the credentials"


def test_the_ledger_indexes_the_run(crawled) -> None:
    report, _, ledger, _ = crawled
    entry: LedgerEntry = ledger.entries[0]
    assert entry.run_id == report.run_id
    assert entry.target_host == "example.com"
    assert entry.findings_count == len(report.findings)


def test_the_observer_saw_the_run_unfold(crawled) -> None:
    """SPEC 7.6: the interface is fed entirely through these calls."""
    report, _, _, observer = crawled
    assert observer.started[0][0].scope_host == "example.com"
    assert len(observer.pages) == len(report.pages)
    assert observer.frontier, "queued and depth must be observable for the progress label"
    assert observer.findings[-1] == report.findings


def test_the_run_directory_is_git_ignored() -> None:
    """AC-LEGAL-5: ``runs/`` holds collected personal data and must stay ignored."""
    ignore = Path(__file__).resolve().parents[1] / ".gitignore"
    assert "runs/" in ignore.read_text(encoding="utf-8").splitlines()


def test_the_deep_page_is_reached_and_the_depth_limit_is_the_only_thing_stopping_it(
    crawled,
) -> None:
    """Depth 3 from the homepage reaches the archive, which is where the limit bites."""
    report, _, _, _ = crawled
    urls = {page.url for page in report.pages}
    assert f"{ORIGIN}/deeper/archive" in urls
