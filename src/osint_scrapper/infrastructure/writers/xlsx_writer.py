"""The XLSX export (SPEC 9.3).

Write-only mode is openpyxl's documented low-memory path. Column widths and
frozen panes are not used: they are not part of a write-only worksheet's
verified surface and the specification makes them optional.
"""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.worksheet._write_only import WriteOnlyWorksheet

from osint_scrapper.domain.report import SiteReport
from osint_scrapper.infrastructure.http.rate_limit import HARD_FLOOR_SECONDS
from osint_scrapper.infrastructure.writers.rows import (
    FINDING_COLUMNS,
    PAGE_COLUMNS,
    compliance_rows,
    finding_rows,
    page_rows,
    run_rows,
)
from osint_scrapper.infrastructure.writers.sanitize import guard

WORKBOOK_FILE_NAME = "report.xlsx"
RUN_SHEET = "Run"
FINDINGS_SHEET = "Findings"
PAGES_SHEET = "Pages"
COMPLIANCE_SHEET = "Compliance"


class XlsxReportWriter:
    """Writes ``report.xlsx`` with exactly four sheets, in the documented order."""

    def __init__(self, hard_floor_seconds: float = HARD_FLOOR_SECONDS) -> None:
        self._hard_floor_seconds = hard_floor_seconds

    @property
    def format_name(self) -> str:
        """The export format this writer produces."""
        return "xlsx"

    def write(self, report: SiteReport, destination_dir: Path) -> tuple[Path, ...]:
        """Write the workbook and return the file created."""
        path = destination_dir / WORKBOOK_FILE_NAME
        workbook = Workbook(write_only=True)

        run_sheet = workbook.create_sheet(title=RUN_SHEET)
        _write_pairs(run_sheet, run_rows(report))

        findings_sheet = workbook.create_sheet(title=FINDINGS_SHEET)
        _write_sheet(findings_sheet, FINDING_COLUMNS, finding_rows(report))

        pages_sheet = workbook.create_sheet(title=PAGES_SHEET)
        _write_sheet(pages_sheet, PAGE_COLUMNS, page_rows(report))

        compliance_sheet = workbook.create_sheet(title=COMPLIANCE_SHEET)
        _write_pairs(compliance_sheet, compliance_rows(report, self._hard_floor_seconds))

        workbook.save(path)
        return (path,)


def _write_pairs(sheet: WriteOnlyWorksheet, rows: Iterable[tuple[str, object]]) -> None:
    _write_header(sheet, ("key", "value"))
    for key, value in rows:
        sheet.append([key, _cell(value)])


def _write_sheet(
    sheet: WriteOnlyWorksheet, columns: Sequence[str], rows: Iterable[tuple[object, ...]]
) -> None:
    _write_header(sheet, columns)
    for row in rows:
        sheet.append([_cell(value) for value in row])


def _write_header(sheet: WriteOnlyWorksheet, columns: Sequence[str]) -> None:
    bold = Font(bold=True)
    header = []
    for column in columns:
        cell = WriteOnlyCell(sheet, value=column)
        cell.font = bold
        header.append(cell)
    sheet.append(header)


def _cell(value: object) -> object:
    """Numbers stay typed; every string is guarded against formula injection."""
    if isinstance(value, bool) or not isinstance(value, str):
        return value
    return guard(value)
